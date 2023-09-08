import openpyxl
import os


os.chdir('C:\\Users\\MZarrilli\\Solmetex\\Solmetex - SQMS Database - General\\1002 - Production Work Order Log')
wb = openpyxl.load_workbook('Work Order Tracker - Shared Master.xlsx')
sheet = wb['Master']
print('enter Part Number')
PartNumber = input()
        #print('enter Lot Number')
        #LotNumber = input()
        #print('enter Requested Quantity')
        #Quantity = input()
        #print('enter Product Description')
        #ProDescript = input()

for row in ws.iter.rows('A{}:A{}'):
    for cell in row:
        cellContent = str(cell.value)
        if cellContent == PartNumber:
            print(cellContent)



#for i in range(1, 500):
 #   cellContent = str(cell.value)
  #  if cellContent == str(PartNumber):
   #     print(cellContent)
        
    # print(i, sheet.cell(column=1).value) 
    # print(i, sheet.cell(row=PartNumber, column=1).value)

